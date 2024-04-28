import tkinter as tk
import openpyxl
from tkinter import filedialog


class WorkDoc:
    def __int__(self):
        self.links_columns_numbers = None
        self.file_path = ''
        self.wb = None
        self.sheet = None
        self.column_count = 0
        self.row_count = 0

    def open_file(self, file_path):
        self.file_path = file_path
        self.wb = openpyxl.load_workbook(file_path)

    def save_doc(self):
        self.wb.save(self.file_path.replace('.xlsx', '-copy.xlsx'))

    def set_sheet(self, sheet_number):
        sheet_names = self.wb.sheetnames
        self.sheet = self.wb[sheet_names[sheet_number]]
        self.column_count = self.sheet.max_column
        self.row_count = self.sheet.max_row

    def get_column_max_count(self):
        return self.column_count

    def get_column_sheet_count(self):
        return self.sheet.max_column

    def get_row_count(self):
        return self.row_count

    def find_column_with_link(self):
        self.links_columns_numbers = []
        for colum in range(2, self.sheet.max_column+1):
            item_link = str(self.sheet.cell(2, colum).value)
            if item_link.find("disk.yandex.ru/d/") >= 0:
                self.links_columns_numbers.append(
                    {
                        'source': colum,
                        'destination': self.column_count + 1
                    }
                )
                self.column_count += 1

        return self.links_columns_numbers

    def write_links_sell(self, row, column, links_str):
        self.sheet.cell(row, column).value = links_str

    def read_sell_value(self, row, column):
        return self.sheet.cell(row, column).value



